"""
Génération du rapport Excel.

Ce n'est pas un export figé : le classeur est un modèle vivant. Les
cellules bleues sont des hypothèses modifiables, tout le reste est
calculé par des formules Excel. Change le prix dans l'onglet Financement
et les mensualités, l'échéancier et le verdict se recalculent.

Sept onglets : Synthèse, Financement, Échéancier, Marché, Contexte,
Checklist, Lexique.

Palette reprise de ui.py, direction papier de géomètre.
"""

import io
from datetime import date

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.marker import DataPoint
from openpyxl.comments import Comment
from openpyxl.drawing.line import LineProperties
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# Palette, sans dièse pour openpyxl
ENCRE = "16303F"
PAPIER = "EDF1F4"
TRAIT = "C3D0D8"
BATI = "1F4E5F"
OR = "E8B33D"
OR_CLAIR = "F0D48A"
FAVORABLE = "2E7D5B"
SANGUINE = "B3402E"
BRUME = "6B8290"
BLANC = "FFFFFF"

POLICE = "Arial"
EUR = '#,##0\\ "€";(#,##0\\ "€");-'
EUR2 = '#,##0.00\\ "€";(#,##0.00\\ "€");-'
PCT = "0.0%"
NB = '#,##0;(#,##0);-'

_bas = Side(style="thin", color=TRAIT)
BORDURE = Border(left=_bas, right=_bas, top=_bas, bottom=_bas)


def _titre(ws, cellule, texte, taille=16):
    ws[cellule] = texte
    ws[cellule].font = Font(name=POLICE, size=taille, bold=True, color=ENCRE)


def _entete(ws, ligne, valeurs, largeur=None):
    for i, v in enumerate(valeurs, start=1):
        c = ws.cell(row=ligne, column=i, value=v)
        c.font = Font(name=POLICE, size=10, bold=True, color=BLANC)
        c.fill = PatternFill("solid", fgColor=ENCRE)
        c.alignment = Alignment(horizontal="center", vertical="center",
                                wrap_text=True)
        c.border = BORDURE
    ws.row_dimensions[ligne].height = 28
    for i, l in enumerate(largeur or [], start=1):
        ws.column_dimensions[get_column_letter(i)].width = l


def _saisie(ws, cellule, valeur, format_=None, commentaire=None):
    """Cellule d'hypothèse : bleu, modifiable par le lecteur."""
    c = ws[cellule]
    c.value = valeur
    c.font = Font(name=POLICE, size=10, color="0000FF")
    c.fill = PatternFill("solid", fgColor="FFFFCC")
    c.border = BORDURE
    if format_:
        c.number_format = format_
    if commentaire:
        c.comment = Comment(commentaire, "Rapport accession")
    return c


def _calcul(ws, cellule, formule, format_=None, gras=False):
    """Cellule calculée : noire, formule Excel."""
    c = ws[cellule]
    c.value = formule
    c.font = Font(name=POLICE, size=10, bold=gras, color=ENCRE)
    c.border = BORDURE
    if format_:
        c.number_format = format_
    return c


def _libelle(ws, cellule, texte, gras=False, indent=0):
    c = ws[cellule]
    c.value = texte
    c.font = Font(name=POLICE, size=10, bold=gras, color=ENCRE)
    c.alignment = Alignment(indent=indent, vertical="center")
    return c


def _bandeau(ws, ligne, texte, detail, favorable):
    fond = FAVORABLE if favorable else SANGUINE
    ws.merge_cells(start_row=ligne, start_column=1, end_row=ligne, end_column=6)
    c = ws.cell(row=ligne, column=1, value=texte)
    c.font = Font(name=POLICE, size=13, bold=True, color=BLANC)
    c.fill = PatternFill("solid", fgColor=fond)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[ligne].height = 30
    ws.merge_cells(start_row=ligne + 1, start_column=1, end_row=ligne + 2,
                   end_column=6)
    d = ws.cell(row=ligne + 1, column=1, value=detail)
    d.font = Font(name=POLICE, size=10, color=ENCRE)
    d.alignment = Alignment(horizontal="left", vertical="top", indent=1,
                            wrap_text=True)
    d.fill = PatternFill("solid", fgColor=PAPIER)


def _legende_couleurs(ws, ligne):
    _libelle(ws, f"A{ligne}", "Légende", gras=True)
    for i, (txt, couleur, police) in enumerate([
        ("Hypothèse modifiable", "FFFFCC", "0000FF"),
        ("Valeur calculée par formule", BLANC, ENCRE),
    ]):
        c = ws.cell(row=ligne + 1 + i, column=1, value=txt)
        c.font = Font(name=POLICE, size=9, color=police)
        c.fill = PatternFill("solid", fgColor=couleur)
        c.border = BORDURE


# ----------------------------------------------------------------------
# Onglet Financement : le modèle vivant
# ----------------------------------------------------------------------

def _feuille_financement(wb, prix, surface, m, cap, dispositif, redevance_m2,
                         baremes):
    """Le modèle vivant. Reproduit exactement la logique de lib.montage :
    l'apport paie d'abord les frais, seul le reliquat réduit l'emprunt."""
    ws = wb.create_sheet("Financement")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 54

    _titre(ws, "A1", "Plan de financement")
    _libelle(ws, "A2", "Les cellules jaunes à texte bleu sont des hypothèses : "
                       "modifie-les, tout le classeur se recalcule.")
    ws["A2"].font = Font(name=POLICE, size=9, italic=True, color=BRUME)

    _libelle(ws, "A4", "HYPOTHÈSES", gras=True)
    taux_notaire = (baremes["frais"]["notaire_ancien_libre"]
                    if dispositif == "Ancien libre"
                    else baremes["frais"]["notaire_neuf_ou_brs"])
    lignes = [
        ("Prix du bien", prix, EUR, "Prix affiché, hors frais."),
        ("Surface habitable (m²)", surface, NB, None),
        ("Apport injecté", m["apport"], EUR,
         "Il paie d'abord les frais. Seul le reliquat réduit l'emprunt. "
         "Attention : trop d'apport bride le PTZ, voir Lexique."),
        ("Quotité PTZ", m["quotite"], PCT,
         "Part du prix finançable à taux zéro, selon la tranche de revenus."),
        ("Prêt Action Logement", m["action_logement"], EUR,
         "30 000 € maximum, 40 000 € en vente HLM."),
        ("Taux Action Logement", baremes["action_logement"]["taux"], PCT, None),
        ("Taux prêt principal", baremes["pret_principal"]["taux_pas_indicatif"],
         PCT, "Taux indicatif constaté été 2026. Remplace-le par ton offre."),
        ("Durée (années)", baremes["pret_principal"]["duree_max_ans"], NB, None),
        ("Taux frais de notaire", taux_notaire, PCT,
         "Environ 3 % en BRS, PSLA et neuf. 7,5 % dans l'ancien libre."),
        ("Taux frais de garantie", baremes["frais"]["garantie_bancaire"], PCT,
         "Caution ou hypothèque. Demande les deux devis."),
        ("Redevance foncière (€/m²/mois)",
         (redevance_m2 if dispositif == "BRS" else 0), EUR2,
         "Plafonnée autour de 1,70 € sur la Métropole de Lyon."),
        ("Taux assurance annuel",
         baremes["pret_principal"]["taux_assurance_annuel_sur_capital"], PCT,
         "La délégation d'assurance fait souvent gagner 5 000 à 15 000 €."),
        ("Différé du PTZ (années)", m["differe_ans"], NB,
         "Période pendant laquelle tu ne rembourses rien sur le PTZ."),
        ("Durée totale du PTZ (années)", m["duree_ptz_ans"] or 25, NB, None),
        ("Capacité mensuelle", cap["disponible"], EUR,
         "35 % des revenus retenus, moins les crédits en cours."),
    ]
    for i, (lib, val, fmt, com) in enumerate(lignes):
        r = 5 + i
        _libelle(ws, f"A{r}", lib, indent=1)
        _saisie(ws, f"B{r}", val, fmt, com)
    ref = {lib: f"$B${5 + i}" for i, (lib, *_) in enumerate(lignes)}
    R = lambda k: ref[k]

    # Frais et coût total
    fr = 5 + len(lignes) + 1
    _libelle(ws, f"A{fr}", "FRAIS ET COÛT TOTAL", gras=True)
    _libelle(ws, f"A{fr + 1}", "Frais de notaire", indent=1)
    _calcul(ws, f"B{fr + 1}",
            f"={R('Prix du bien')}*{R('Taux frais de notaire')}", EUR)
    _libelle(ws, f"A{fr + 2}", "Frais de garantie", indent=1)
    _calcul(ws, f"B{fr + 2}",
            f"={R('Prix du bien')}*{R('Taux frais de garantie')}", EUR)
    _libelle(ws, f"A{fr + 3}", "Coût total de l'opération", gras=True)
    _calcul(ws, f"B{fr + 3}",
            f"={R('Prix du bien')}+B{fr + 1}+B{fr + 2}", EUR, gras=True)

    # Plan de financement
    d = fr + 5
    _libelle(ws, f"A{d}", "PLAN DE FINANCEMENT", gras=True)
    _entete(ws, d + 1, ["Ressource", "Montant", "Part du prix"])
    r_ptz, r_al, r_ap, r_pr = d + 2, d + 3, d + 4, d + 5
    _libelle(ws, f"A{r_ptz}", "Prêt à taux zéro", indent=1)
    _calcul(ws, f"B{r_ptz}", f"={R('Prix du bien')}*{R('Quotité PTZ')}", EUR)
    _libelle(ws, f"A{r_al}", "Prêt Action Logement à 1 %", indent=1)
    _calcul(ws, f"B{r_al}", f"={R('Prêt Action Logement')}", EUR)
    _libelle(ws, f"A{r_ap}", "Apport affecté au prix", indent=1)
    _calcul(ws, f"B{r_ap}",
            f"=MAX(0,{R('Apport injecté')}-B{fr + 1}-B{fr + 2})", EUR)
    _libelle(ws, f"A{r_pr}", "Prêt principal (le solde)", indent=1)
    _calcul(ws, f"B{r_pr}",
            f"=MAX(0,{R('Prix du bien')}-B{r_ptz}-B{r_al}-B{r_ap})", EUR)
    for r in (r_ptz, r_al, r_ap, r_pr):
        _calcul(ws, f"C{r}", f"=IFERROR(B{r}/{R('Prix du bien')},0)", PCT)
    r_tot = r_pr + 1
    _libelle(ws, f"A{r_tot}", "Total des ressources", gras=True)
    _calcul(ws, f"B{r_tot}", f"=SUM(B{r_ptz}:B{r_pr})", EUR, gras=True)
    _calcul(ws, f"C{r_tot}", f"=IFERROR(B{r_tot}/{R('Prix du bien')},0)", PCT,
            gras=True)
    r_ctrl = r_tot + 1
    _libelle(ws, f"A{r_ctrl}", "Contrôle, total contre prix", indent=1)
    _calcul(ws, f"B{r_ctrl}",
            f'=IF(ABS(B{r_tot}-{R("Prix du bien")})<1,"cohérent",'
            f'"écart à vérifier")')
    r_regle = r_ctrl + 1
    _libelle(ws, f"A{r_regle}", "Règle du PTZ contre les autres prêts", indent=1)
    _calcul(ws, f"B{r_regle}",
            f'=IF(B{r_ptz}<=IF({R("Quotité PTZ")}>=0.5,1.25,1)*'
            f'(B{r_al}+B{r_pr}),"respectée","PTZ bridé : réduis ton apport")')
    ws[f"C{r_regle}"] = ("Le PTZ ne peut dépasser le total des autres prêts, "
                         "sauf à quotité 50 % où il peut les dépasser de 25 %.")
    ws[f"C{r_regle}"].font = Font(name=POLICE, size=9, color=BRUME)
    ws[f"C{r_regle}"].alignment = Alignment(wrap_text=True, vertical="center")
    r_grat = r_regle + 1
    _libelle(ws, f"A{r_grat}", "Part sans intérêts ou à 1 %", gras=True)
    _calcul(ws, f"B{r_grat}",
            f"=IFERROR((B{r_ptz}+B{r_al})/{R('Prix du bien')},0)", PCT,
            gras=True)

    # Mensualités
    e = r_grat + 2
    _libelle(ws, f"A{e}", "MENSUALITÉS", gras=True)
    _entete(ws, e + 1, ["Composante", "Par mois", "Commentaire"])
    mens = [
        ("Prêt Action Logement",
         f"=IFERROR(-PMT({R('Taux Action Logement')}/12,"
         f"{R('Durée (années)')}*12,B{r_al}),0)",
         "Amortissement à 1 % sur la durée retenue."),
        ("Prêt principal",
         f"=IFERROR(-PMT({R('Taux prêt principal')}/12,"
         f"{R('Durée (années)')}*12,B{r_pr}),0)",
         "Amortissement au taux du marché."),
        ("Assurance emprunteur",
         f"=(B{r_ptz}+B{r_al}+B{r_pr})*{R('Taux assurance annuel')}/12",
         "Poste le plus négociable : compare en délégation."),
        ("Redevance foncière",
         f"={R('Surface habitable (m²)')}*"
         f"{R('Redevance foncière (€/m²/mois)')}",
         "En BRS uniquement. Charge à vie, ne construit aucun capital."),
    ]
    for i, (lib, f, com) in enumerate(mens):
        r = e + 2 + i
        _libelle(ws, f"A{r}", lib, indent=1)
        _calcul(ws, f"B{r}", f, EUR2)
        _libelle(ws, f"C{r}", com)
        ws[f"C{r}"].font = Font(name=POLICE, size=9, color=BRUME)
        ws[f"C{r}"].alignment = Alignment(wrap_text=True, vertical="center")
    r_p1 = e + 2 + len(mens)
    _libelle(ws, f"A{r_p1}", "PHASE 1, pendant le différé", gras=True)
    _calcul(ws, f"B{r_p1}", f"=SUM(B{e + 2}:B{r_p1 - 1})", EUR2, gras=True)
    _libelle(ws, f"C{r_p1}", "Ce que tu paies les premières années.")
    ws[f"C{r_p1}"].font = Font(name=POLICE, size=9, color=BRUME)
    r_mptz = r_p1 + 1
    _libelle(ws, f"A{r_mptz}", "Amortissement du PTZ après le différé",
             indent=1)
    _calcul(ws, f"B{r_mptz}",
            f"=IFERROR(B{r_ptz}/(MAX({R('Durée totale du PTZ (années)')}-"
            f"{R('Différé du PTZ (années)')},1)*12),0)", EUR2)
    r_p2 = r_mptz + 1
    _libelle(ws, f"A{r_p2}", "PHASE 2, après le différé", gras=True)
    _calcul(ws, f"B{r_p2}", f"=B{r_p1}+B{r_mptz}", EUR2, gras=True)
    ws[f"B{r_p2}"].fill = PatternFill("solid", fgColor=OR_CLAIR)
    _libelle(ws, f"C{r_p2}",
             "C'est CETTE échéance que la banque teste, pas la phase 1.")
    ws[f"C{r_p2}"].font = Font(name=POLICE, size=9, bold=True, color=SANGUINE)
    ws[f"C{r_p2}"].alignment = Alignment(wrap_text=True, vertical="center")

    r_marge = r_p2 + 2
    _libelle(ws, f"A{r_marge}", "Marge sur ta capacité", gras=True)
    _calcul(ws, f"B{r_marge}", f"={R('Capacité mensuelle')}-B{r_p2}", EUR2,
            gras=True)
    r_verdict = r_marge + 1
    _libelle(ws, f"A{r_verdict}", "Verdict", gras=True)
    _calcul(ws, f"B{r_verdict}",
            f'=IF(B{r_marge}>=0,"Finançable",'
            f'"Dépassement en phase 2")', gras=True)

    graphe = BarChart()
    graphe.type = "col"
    graphe.title = "Répartition du financement"
    graphe.y_axis.numFmt = EUR
    graphe.y_axis.title = "Euros"
    graphe.add_data(Reference(ws, min_col=2, min_row=d + 1, max_row=r_pr),
                    titles_from_data=True)
    graphe.set_categories(Reference(ws, min_col=1, min_row=d + 2, max_row=r_pr))
    graphe.height, graphe.width = 8.5, 17
    graphe.legend = None
    serie = graphe.series[0]
    serie.graphicalProperties.solidFill = BATI
    for idx, couleur in enumerate([OR, OR_CLAIR, BRUME, BATI]):
        pt = DataPoint(idx=idx)
        pt.graphicalProperties.solidFill = couleur
        pt.graphicalProperties.line = LineProperties(solidFill=ENCRE)
        serie.data_points.append(pt)
    ws.add_chart(graphe, f"E{d}")

    _legende_couleurs(ws, r_verdict + 2)
    return {
        "phase1": f"Financement!$B${r_p1}", "phase2": f"Financement!$B${r_p2}",
        "differe": f"Financement!{R('Différé du PTZ (années)')}",
        "capacite": f"Financement!{R('Capacité mensuelle')}",
        "ptz": f"Financement!$B${r_ptz}", "al": f"Financement!$B${r_al}",
        "principal": f"Financement!$B${r_pr}",
        "verdict": f"Financement!$B${r_verdict}",
        "marge": f"Financement!$B${r_marge}",
        "gratuit": f"Financement!$B${r_grat}",
        "cout_total": f"Financement!$B${fr + 3}",
        "prix": f"Financement!{R('Prix du bien')}",
        "surface": f"Financement!{R('Surface habitable (m²)')}",
    }


# ----------------------------------------------------------------------
# Onglet Échéancier
# ----------------------------------------------------------------------

def _feuille_echeancier(wb, liens, duree=25):
    ws = wb.create_sheet("Échéancier")
    ws.sheet_view.showGridLines = False
    _titre(ws, "A1", "Ton échéancier sur 25 ans")
    _libelle(ws, "A2", "Le saut correspond à la fin du différé du prêt à taux "
                       "zéro. La ligne plate est ta capacité.")
    ws["A2"].font = Font(name=POLICE, size=9, italic=True, color=BRUME)
    _entete(ws, 4, ["Année", "Mensualité tout compris", "Ta capacité",
                    "Marge"], [10, 24, 16, 14])
    for i in range(1, duree + 1):
        r = 4 + i
        _calcul(ws, f"A{r}", i, NB)
        _calcul(ws, f"B{r}",
                f"=IF(A{r}<={liens['differe']},{liens['phase1']},"
                f"{liens['phase2']})", EUR)
        _calcul(ws, f"C{r}", f"={liens['capacite']}", EUR)
        _calcul(ws, f"D{r}", f"=C{r}-B{r}", EUR)
        ws[f"D{r}"].font = Font(name=POLICE, size=10, color=ENCRE)

    graphe = LineChart()
    graphe.title = "Mensualité et capacité, année par année"
    graphe.y_axis.numFmt = EUR
    graphe.y_axis.title = "Euros par mois"
    graphe.x_axis.title = "Année"
    donnees = Reference(ws, min_col=2, max_col=3, min_row=4, max_row=4 + duree)
    graphe.add_data(donnees, titles_from_data=True)
    graphe.set_categories(Reference(ws, min_col=1, min_row=5, max_row=4 + duree))
    graphe.height, graphe.width = 9, 18
    graphe.series[0].graphicalProperties.line.solidFill = BATI
    graphe.series[0].graphicalProperties.line.width = 28000
    graphe.series[1].graphicalProperties.line.solidFill = SANGUINE
    graphe.series[1].graphicalProperties.line.dashStyle = "dash"
    for s in graphe.series:
        s.smooth = False
    ws.add_chart(graphe, "F4")


# ----------------------------------------------------------------------
# Onglet Synthèse
# ----------------------------------------------------------------------

def _feuille_synthese(wb, contexte, liens, dec):
    ws = wb.create_sheet("Synthèse", 0)
    ws.sheet_view.showGridLines = False
    for col, w in zip("ABCDEF", [30, 20, 20, 18, 18, 18]):
        ws.column_dimensions[col].width = w

    _titre(ws, "A1", "Clé de Sol, fiche de chiffrage", 18)
    _libelle(ws, "A2", f"Éditée le {date.today().strftime('%d/%m/%Y')} par "
                       f"ton outil personnel de recherche.")
    ws["A2"].font = Font(name=POLICE, size=9, italic=True, color=BRUME)

    _bandeau(ws, 4, contexte["verdict_titre"], contexte["verdict_detail"],
             contexte["favorable"])

    _libelle(ws, "A8", "LE BIEN", gras=True)
    infos = [("Adresse", contexte.get("adresse") or "non renseignée"),
             ("Commune", contexte.get("commune") or "non renseignée"),
             ("Dispositif", contexte["dispositif"]),
             ("Prix affiché", None), ("Surface habitable", None),
             ("Prix au m²", None)]
    for i, (lib, _) in enumerate(infos):
        r = 9 + i
        _libelle(ws, f"A{r}", lib, indent=1)
    _libelle(ws, "B9", infos[0][1])
    _libelle(ws, "B10", infos[1][1])
    _libelle(ws, "B11", infos[2][1])
    _calcul(ws, "B12", f"={liens['prix']}", EUR)
    _calcul(ws, "B13", f"={liens['surface']}", NB)
    _calcul(ws, "B14", f"=IFERROR({liens['prix']}/{liens['surface']},0)", EUR)

    _libelle(ws, "A16", "LE FINANCEMENT", gras=True)
    for i, (lib, f, fmt) in enumerate([
        ("Prêt à taux zéro", liens["ptz"], EUR),
        ("Prêt Action Logement à 1 %", liens["al"], EUR),
        ("Prêt principal", liens["principal"], EUR),
        ("Part sans intérêts ou à 1 %", liens["gratuit"], PCT),
    ]):
        r = 17 + i
        _libelle(ws, f"A{r}", lib, indent=1)
        _calcul(ws, f"B{r}", f"={f}", fmt)

    _libelle(ws, "A22", "LES MENSUALITÉS", gras=True)
    for i, (lib, f) in enumerate([
        ("Phase 1, pendant le différé", liens["phase1"]),
        ("Phase 2, après le différé", liens["phase2"]),
        ("Ta capacité mensuelle", liens["capacite"]),
        ("Marge au point le plus haut", liens["marge"]),
    ]):
        r = 23 + i
        _libelle(ws, f"A{r}", lib, indent=1)
        _calcul(ws, f"B{r}", f"={f}", EUR, gras=(i in (1, 3)))
    ws["B24"].fill = PatternFill("solid", fgColor=OR_CLAIR)

    if dec and dec.get("ok"):
        _libelle(ws, "A28", "LE MARCHÉ LOCAL", gras=True)
        for i, (lib, val, fmt) in enumerate([
            ("Médiane du secteur (€/m²)", round(dec["mediane_m2"]), EUR),
            ("Décote du bien", dec["decote_pct"] / 100, PCT),
            ("Ventes comparables analysées", dec["nb_ventes"], NB),
        ]):
            r = 29 + i
            _libelle(ws, f"A{r}", lib, indent=1)
            c = _calcul(ws, f"B{r}", val, fmt)
            c.font = Font(name=POLICE, size=10, color=ENCRE)

    r = 34
    _libelle(ws, f"A{r}", "COMMENT LIRE CE CLASSEUR", gras=True)
    for i, txt in enumerate([
        "Onglet Financement : le modèle. Les cellules jaunes sont modifiables, "
        "tout le reste se recalcule.",
        "Onglet Échéancier : le saut de mensualité à la fin du différé, face à "
        "ta capacité.",
        "Onglet Marché : les ventes réelles du quartier qui servent au calcul "
        "de décote.",
        "Onglet Contexte : énergie, risques, écoles, équipements.",
        "Onglet Checklist : les 25 points de la recherche aux clés.",
        "Onglet Lexique : chaque terme technique expliqué.",
    ]):
        _libelle(ws, f"A{r + 1 + i}", "- " + txt, indent=1)
        ws[f"A{r + 1 + i}"].font = Font(name=POLICE, size=9, color=ENCRE)

    _libelle(ws, "A42", "Estimations produites par un outil personnel. Elles ne "
                        "remplacent ni un courtier, ni un notaire, ni l'ADIL.")
    ws["A42"].font = Font(name=POLICE, size=9, italic=True, color=SANGUINE)


# ----------------------------------------------------------------------
# Onglets de données
# ----------------------------------------------------------------------

def _feuille_marche(wb, ventes, dec):
    ws = wb.create_sheet("Marché")
    ws.sheet_view.showGridLines = False
    _titre(ws, "A1", "Les ventes réelles du quartier")
    _libelle(ws, "A2", "Source : fichier public des demandes de valeurs "
                       "foncières, enregistrées par les notaires.")
    ws["A2"].font = Font(name=POLICE, size=9, italic=True, color=BRUME)
    if not ventes:
        _libelle(ws, "A4", "Aucune vente comparable trouvée dans le rayon.")
        return
    _entete(ws, 4, ["Date", "Type", "Surface m²", "Prix", "Prix au m²"],
            [14, 22, 12, 16, 14])
    for i, v in enumerate(sorted(ventes, key=lambda x: str(x.get("date") or ""),
                                 reverse=True)[:300]):
        r = 5 + i
        _calcul(ws, f"A{r}", str(v.get("date") or ""))
        _calcul(ws, f"B{r}", str(v.get("type") or ""))
        _calcul(ws, f"C{r}", round(v["surface"]), NB)
        _calcul(ws, f"D{r}", round(v["prix"]), EUR)
        _calcul(ws, f"E{r}", f"=IFERROR(D{r}/C{r},0)", EUR)
    fin = 4 + min(len(ventes), 300)
    s = fin + 2
    _libelle(ws, f"A{s}", "STATISTIQUES", gras=True)
    for i, (lib, f) in enumerate([
        ("Nombre de ventes", f"=COUNT(E5:E{fin})"),
        ("Médiane du prix au m²", f"=MEDIAN(E5:E{fin})"),
        ("Prix au m² le plus bas", f"=MIN(E5:E{fin})"),
        ("Prix au m² le plus haut", f"=MAX(E5:E{fin})"),
    ]):
        r = s + 1 + i
        _libelle(ws, f"A{r}", lib, indent=1)
        _calcul(ws, f"B{r}", f, NB if i == 0 else EUR, gras=(i == 1))


def _feuille_contexte(wb, contexte):
    ws = wb.create_sheet("Contexte")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 46
    _titre(ws, "A1", "Le contexte de l'adresse")
    ligne = 3
    for section, elements in contexte.get("blocs", []):
        _libelle(ws, f"A{ligne}", section.upper(), gras=True)
        ligne += 1
        for cle, valeur in elements:
            _libelle(ws, f"A{ligne}", str(cle), indent=1)
            _libelle(ws, f"B{ligne}", str(valeur))
            ws[f"B{ligne}"].border = BORDURE
            ligne += 1
        ligne += 1
    if ligne == 3:
        _libelle(ws, "A3", "Aucune donnée de contexte disponible.")


def _feuille_checklist(wb, checklist):
    ws = wb.create_sheet("Checklist")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 78
    _titre(ws, "A1", "De la recherche aux clés")
    _libelle(ws, "A2", "Coche la colonne de gauche au fur et à mesure.")
    ws["A2"].font = Font(name=POLICE, size=9, italic=True, color=BRUME)
    ligne = 4
    for phase, elements in checklist:
        ws.merge_cells(start_row=ligne, start_column=1, end_row=ligne,
                       end_column=2)
        c = ws.cell(row=ligne, column=1, value=phase)
        c.font = Font(name=POLICE, size=11, bold=True, color=BLANC)
        c.fill = PatternFill("solid", fgColor=BATI)
        c.alignment = Alignment(indent=1, vertical="center")
        ws.row_dimensions[ligne].height = 22
        ligne += 1
        for e in elements:
            case = ws.cell(row=ligne, column=1, value="")
            case.fill = PatternFill("solid", fgColor="FFFFCC")
            case.border = BORDURE
            case.alignment = Alignment(horizontal="center")
            t = ws.cell(row=ligne, column=2, value=e)
            t.font = Font(name=POLICE, size=10, color=ENCRE)
            t.alignment = Alignment(wrap_text=True, vertical="center", indent=1)
            t.border = BORDURE
            ws.row_dimensions[ligne].height = 24
            ligne += 1
        ligne += 1


def _feuille_lexique(wb, glossaire, ordre):
    ws = wb.create_sheet("Lexique")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 96
    _titre(ws, "A1", "Chaque terme, en français")
    _entete(ws, 3, ["Terme", "Ce que ça veut dire"])
    for i, cle in enumerate(ordre):
        if cle not in glossaire:
            continue
        r = 4 + i
        c = ws.cell(row=r, column=1, value=cle.replace("_", " ").upper())
        c.font = Font(name=POLICE, size=10, bold=True, color=BATI)
        c.alignment = Alignment(vertical="top", indent=1)
        c.border = BORDURE
        t = ws.cell(row=r, column=2, value=glossaire[cle])
        t.font = Font(name=POLICE, size=10, color=ENCRE)
        t.alignment = Alignment(wrap_text=True, vertical="top", indent=1)
        t.border = BORDURE
        ws.row_dimensions[r].height = 58


# ----------------------------------------------------------------------
# Point d'entrée
# ----------------------------------------------------------------------

def construire(prix, surface, dispositif, m, cap, glossaire, checklist,
               baremes, redevance_m2=1.7, adresse=None, commune=None,
               dec=None, ventes=None, blocs_contexte=None) -> bytes:
    """Assemble le classeur et renvoie les octets, prêts pour le téléchargement."""
    wb = Workbook()
    wb.remove(wb.active)

    liens = _feuille_financement(wb, prix, surface, m, cap, dispositif,
                                 redevance_m2, baremes)
    marge = cap["disponible"] - m["phase2"]
    favorable = marge >= 0
    contexte = {
        "adresse": adresse, "commune": commune, "dispositif": dispositif,
        "favorable": favorable,
        "verdict_titre": (
            f"Finançable, marge de {round(marge)} € par mois" if favorable
            else f"Dépassement de {round(-marge)} € par mois en phase 2"),
        "verdict_detail": (
            "La phase 2 est l'échéance la plus haute de ton crédit, celle qui "
            "démarre à la fin du différé du prêt à taux zéro. C'est elle que la "
            "banque teste, et c'est sur elle qu'est calculée cette marge."
            if favorable else
            "Trois leviers : rembourser un crédit en cours, demander le "
            "lissage des prêts à la banque, ou réduire le prêt à taux zéro "
            "pour aplatir le profil de l'échéancier."),
        "blocs": blocs_contexte or [],
    }
    _feuille_synthese(wb, contexte, liens, dec)
    _feuille_echeancier(wb, liens)
    _feuille_marche(wb, ventes or [], dec)
    _feuille_contexte(wb, contexte)
    _feuille_checklist(wb, checklist)
    _feuille_lexique(wb, glossaire, [
        "ptz", "quotite", "differe", "tranche", "phase2",
        "regle_ptz_autres_prets", "action_logement", "pas", "hcsf", "capacite",
        "brs", "psla", "redevance", "rfr", "zone", "occupants", "dvf",
        "decote", "dpe", "georisques", "argile", "radon", "ips", "parcelle",
        "gpu",
    ])

    tampon = io.BytesIO()
    wb.save(tampon)
    return tampon.getvalue()
